from fastapi import APIRouter, Depends, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse

from sqlalchemy.orm import Session, selectinload, undefer
from sqlalchemy import func as func_count
from typing import Optional
from datetime import date
from .. import models, auth as auth_module
from ..templates_config import templates
from ..database import get_db
from ..models import CondicionBoleta


def _parse_zona_id(zona_id_str: Optional[str]) -> Optional[int]:
    """Convierte zona_id del form a int, ignorando el valor '__nueva__'."""
    if not zona_id_str or zona_id_str == "__nueva__":
        return None
    try:
        return int(zona_id_str)
    except ValueError:
        return None

router = APIRouter(prefix="/compradores", tags=["compradores"])



@router.get("/", response_class=HTMLResponse)
async def listar(request: Request, db: Session = Depends(get_db),
                 q: str = "", pata: str = "", zona: str = "",
                 sin_cob: str = "", cob: str = "",
                 vend: str = "", cont: str = "",
                 mes: str = "", anio: str = ""):
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')

    from sqlalchemy import or_ as sql_or

    # Base query: compradores que tengan al menos una boleta
    query = db.query(models.Comprador).join(models.Zona, isouter=True)
    if q:
        query = query.filter(models.Comprador.apellido_nombre.ilike(f"%{q}%"))

    # ── Filtros que dependen de la tabla Boleta ──────────────────────────────
    # IMPORTANTE: la tabla Boleta se JOINea UNA SOLA VEZ. Si se combinaban dos
    # filtros que la requerían (p. ej. vendedor + cuotas/contado), antes se hacía
    # un segundo .join(Boleta) y SQL fallaba con "table name specified more than
    # once" → Internal Server Error. Ahora se joinea una vez y se acumulan los
    # filtros sobre ese mismo join.
    _cob_id = None
    if cob:
        try:
            _cob_id = int(cob)
        except (TypeError, ValueError):
            _cob_id = None
    _vend_id = None
    if vend:
        try:
            _vend_id = int(vend)
        except (TypeError, ValueError):
            _vend_id = None

    _mes = None
    if mes:
        try:
            _mes = int(mes)
            if not (1 <= _mes <= 12):
                _mes = None
        except (TypeError, ValueError):
            _mes = None
    _anio = None
    if anio:
        try:
            _anio = int(anio)
        except (TypeError, ValueError):
            _anio = None

    from sqlalchemy import func as sqlfunc

    needs_boleta = bool(pata) or sin_cob in ("1", "true", "yes") or \
        _cob_id is not None or _vend_id is not None or cont in ("contado", "cuotas") or \
        _mes is not None or _anio is not None

    if needs_boleta:
        query = query.join(models.Boleta,
                           models.Boleta.comprador_id == models.Comprador.id)

    if pata:
        query = (query
                 .join(models.Talonera, models.Talonera.id == models.Boleta.talonera_id)
                 .filter(models.Talonera.nombre == pata))
    # Filtro: solo socios sin cobrador en alguna boleta. Se excluyen las boletas
    # con todas las cuotas pagas (al contado / finalizadas): no necesitan cobrador.
    if sin_cob in ("1", "true", "yes"):
        query = query.filter(
            models.Boleta.cobrador_id.is_(None),
            models.Boleta.cuotas_pagadas < models.Boleta.cuotas_pactadas,
        )
    # Filtro: socios cuyo cobrador actual es X.
    # Excluye contados: aunque la boleta conserve cobrador_id (bug viejo de
    # auto-asignación), un contado no lo cobra el cobrador y no debe aparecer
    # en su listado. Mismo criterio que el badge CONTADO del template:
    # numero_especial / numero_especial_2 asignado, o anticipo_total
    # (cuotas_pactadas>0 y cuotas_anticipadas >= cuotas_pactadas).
    if _cob_id is not None:
        query = query.filter(
            models.Boleta.cobrador_id == _cob_id,
            models.Boleta.numero_especial.is_(None),
            models.Boleta.numero_especial_2.is_(None),
            sql_or(
                func_count.coalesce(models.Boleta.cuotas_pactadas, 0) == 0,
                func_count.coalesce(models.Boleta.cuotas_anticipadas, 0)
                    < func_count.coalesce(models.Boleta.cuotas_pactadas, 0),
            ),
        )
    # Filtro: socios cuyo vendedor (en alguna boleta) es X
    if _vend_id is not None:
        query = query.filter(models.Boleta.vendedor_id == _vend_id)
    # Filtro: modalidad cuotas vs contado.
    # contado = boleta con numero_especial(_2) asignado.
    # cuotas  = boleta en curso (cuotas_pagadas < cuotas_pactadas).
    if cont == "contado":
        query = query.filter(sql_or(
            models.Boleta.numero_especial.isnot(None),
            models.Boleta.numero_especial_2.isnot(None),
        ))
    elif cont == "cuotas":
        query = query.filter(models.Boleta.cuotas_pagadas < models.Boleta.cuotas_pactadas)
    # Filtro por mes / año de la fecha de compra (fecha_venta)
    if _mes is not None:
        query = query.filter(sqlfunc.extract('month', models.Boleta.fecha_venta) == _mes)
    if _anio is not None:
        query = query.filter(sqlfunc.extract('year', models.Boleta.fecha_venta) == _anio)

    if needs_boleta:
        query = query.distinct()
    if zona:
        # Filtro por zona — acepta id numérico o nombre exacto
        try:
            zona_id_int = int(zona)
            query = query.filter(models.Comprador.zona_id == zona_id_int)
        except (TypeError, ValueError):
            query = query.filter(models.Zona.nombre == zona)
    # undefer(numero_especial_2): el template (compradores.html, badge CONTADO)
    # lee esa columna por cada boleta. Al estar declarada `deferred` en models.py,
    # cada acceso dispara un SELECT propio → con 600 socios eran 600 queries extra
    # por request (en Railway, 600 round-trips de red = varios segundos). Pidiéndola
    # en el SELECT del selectinload el costo pasa a cero.
    compradores_raw = query.options(
        selectinload(models.Comprador.boletas).undefer(models.Boleta.numero_especial_2),
        selectinload(models.Comprador.boletas).selectinload(models.Boleta.talonera)
    ).order_by(models.Comprador.apellido_nombre).all()
    # Ordenar por PATA (multiplicador) primero, luego por número de boleta
    compradores = sorted(
        compradores_raw,
        key=lambda c: (
            c.boletas[0].talonera.multiplicador if c.boletas and c.boletas[0].talonera else 999,
            c.boletas[0].numero_principal if c.boletas else 999999
        )
    )

    # Conteos por talonera para los tabs. También traemos multiplicador para
    # calcular el total "Todas" ponderado (PATA 1 ×1, PATA 2 ×2, ...). Las
    # pestañas individuales siguen mostrando el conteo literal de socios.
    # Los tabs reflejan el subset filtrado por cob/vend/zona/cont/q/sin_cob
    # (excepto pata — porque las pestañas SON el selector de PATA).
    tabs_query = (
        db.query(
            models.Talonera.nombre,
            sqlfunc.count(models.Comprador.id.distinct()),
            models.Talonera.multiplicador,
        )
        .join(models.Boleta, models.Boleta.talonera_id == models.Talonera.id)
        .join(models.Comprador, models.Comprador.id == models.Boleta.comprador_id)
    )
    if q:
        tabs_query = tabs_query.filter(models.Comprador.apellido_nombre.ilike(f"%{q}%"))
    if sin_cob in ("1", "true", "yes"):
        tabs_query = tabs_query.filter(
            models.Boleta.cobrador_id.is_(None),
            models.Boleta.cuotas_pagadas < models.Boleta.cuotas_pactadas,
        )
    if cob:
        try:
            tabs_query = tabs_query.filter(models.Boleta.cobrador_id == int(cob))
        except (TypeError, ValueError):
            pass
    if vend:
        try:
            tabs_query = tabs_query.filter(models.Boleta.vendedor_id == int(vend))
        except (TypeError, ValueError):
            pass
    if cont == "contado":
        tabs_query = tabs_query.filter(sql_or(
            models.Boleta.numero_especial.isnot(None),
            models.Boleta.numero_especial_2.isnot(None),
        ))
    elif cont == "cuotas":
        tabs_query = tabs_query.filter(
            models.Boleta.cuotas_pagadas < models.Boleta.cuotas_pactadas,
        )
    if _mes is not None:
        tabs_query = tabs_query.filter(sqlfunc.extract('month', models.Boleta.fecha_venta) == _mes)
    if _anio is not None:
        tabs_query = tabs_query.filter(sqlfunc.extract('year', models.Boleta.fecha_venta) == _anio)
    if zona:
        try:
            zid = int(zona)
            tabs_query = tabs_query.filter(models.Comprador.zona_id == zid)
        except (TypeError, ValueError):
            tabs_query = (tabs_query
                .join(models.Zona, models.Zona.id == models.Comprador.zona_id)
                .filter(models.Zona.nombre == zona))
    taloneras_raw = (
        tabs_query
        .group_by(models.Talonera.nombre, models.Talonera.multiplicador)
        .order_by(models.Talonera.nombre)
        .all()
    )
    tabs = [
        {"nombre": t[0], "total": t[1], "multiplicador": float(t[2] or 1.0)}
        for t in taloneras_raw
    ]
    # Total ponderado: 39 PATA1×1 + 9 PATA2×2 + 1 PATA3×3 + 1 PATA4×4 + 3 PATA8×8
    # = 39 + 18 + 3 + 4 + 24 = 88. Con PATA 0 (×0.67) el total puede tener decimales —
    # redondeo a entero al mostrar (decisión 11/05/2026).
    total_compradores = round(sum(t["total"] * t["multiplicador"] for t in tabs))

    # Cantidad de socios sin vendedor/cobrador (para mostrar alertas en el template).
    # Para el cobrador: excluimos boletas ya totalmente pagas (al contado / finalizadas)
    # porque esas no necesitan cobrador — ya no hay cuotas pendientes de cobrar.
    sin_vendedor = sum(1 for c in compradores if c.boletas and not c.boletas[0].vendedor_id)
    sin_cobrador = sum(
        1 for c in compradores
        if c.boletas
        and not c.boletas[0].cobrador_id
        and (c.boletas[0].cuotas_pagadas or 0) < (c.boletas[0].cuotas_pactadas or 0)
    )

    # Contado liquidado por vendedor pero todavía sin cargar en sistema:
    # son los LiquidacionContadoItem cuyo número aún no fue asignado
    # a ninguna Boleta con comprador (via numero_especial o numero_especial_2).
    # Solo necesitamos (talonera_id, numero) de cada LCI: seleccionar esas
    # columnas evita materializar objetos ORM completos.
    lci_all = db.query(
        models.LiquidacionContadoItem.talonera_id,
        models.LiquidacionContadoItem.numero,
    ).all()
    if lci_all:
        # Traer SOLO las columnas necesarias como tuplas (una query cada set).
        # Importante: numero_especial_2 / talonera_especial_2_id son columnas
        # `deferred` — recorrer objetos Boleta y accederlas dispara una consulta
        # extra por fila (N+1). Pedirlas explícitas en el SELECT lo evita.
        asignados_slot1 = set(
            db.query(
                models.Boleta.talonera_especial_id,
                models.Boleta.numero_especial,
            ).filter(
                models.Boleta.comprador_id.isnot(None),
                models.Boleta.numero_especial.isnot(None),
            ).all()
        )
        asignados_slot2 = set(
            db.query(
                models.Boleta.talonera_especial_2_id,
                models.Boleta.numero_especial_2,
            ).filter(
                models.Boleta.comprador_id.isnot(None),
                models.Boleta.numero_especial_2.isnot(None),
            ).all()
        )
        contado_sin_cargar = sum(
            1 for (t_id, num) in lci_all
            if (t_id, num) not in asignados_slot1
            and (t_id, num) not in asignados_slot2
        )
    else:
        contado_sin_cargar = 0

    zonas = db.query(models.Zona).order_by(models.Zona.nombre).all()
    vendedores = db.query(models.Vendedor).filter(models.Vendedor.activo == True).order_by(models.Vendedor.nombre).all()
    cobradores = db.query(models.Cobrador).filter(models.Cobrador.activo == True).order_by(models.Cobrador.nombre).all()

    # Años disponibles según fecha de compra (para el filtro Año)
    anios_raw = (
        db.query(sqlfunc.extract('year', models.Boleta.fecha_venta))
        .filter(models.Boleta.fecha_venta.isnot(None))
        .distinct()
        .all()
    )
    anios = sorted({int(a[0]) for a in anios_raw if a[0] is not None}, reverse=True)

    # Contador de boletas/socios que coinciden con el filtro actual
    total_filtrado = len(compradores)
    # Contador ponderado por número (X1×1, X2×2, X0×0.67, ...) — redondeado a entero
    total_filtrado_pond = round(sum(
        (c.boletas[0].talonera.multiplicador
         if c.boletas and c.boletas[0].talonera and c.boletas[0].talonera.multiplicador
         else 1.0)
        for c in compradores
    ))

    return templates.TemplateResponse(request, "compradores.html", {
        "user": user,
        "compradores": compradores,
        "zonas": zonas,
        "vendedores": vendedores,
        "cobradores": cobradores,
        "q": q,
        "pata": pata,
        "tabs": tabs,
        "total_compradores": total_compradores,
        "sin_vendedor": sin_vendedor,
        "sin_cobrador": sin_cobrador,
        "contado_sin_cargar": contado_sin_cargar,
        "filtro_sin_cob": sin_cob in ("1", "true", "yes"),
        "filtro_cob": cob,
        "filtro_vend": vend,
        "filtro_zona": zona,
        "filtro_cont": cont if cont in ("cuotas", "contado") else "",
        "filtro_mes": str(_mes) if _mes is not None else "",
        "filtro_anio": str(_anio) if _anio is not None else "",
        "anios": anios,
        "total_filtrado": total_filtrado,
        "total_filtrado_pond": total_filtrado_pond,
    })


@router.get("/contado-sin-cargar-lista")
async def contado_sin_cargar_lista(request: Request, db: Session = Depends(get_db)):
    """Devuelve los LiquidacionContadoItem que el vendedor declaró como contado
    pero cuyo número aún no fue asignado a ningún Comprador (numero_especial / numero_especial_2).
    Útil para saber qué socios faltan cargar en sistema después de una liquidación.
    """
    await auth_module.require_user(request, db)

    lci_all = (
        db.query(
            models.LiquidacionContadoItem,
            models.LiquidacionVendedor,
            models.Talonera,
            models.Vendedor,
        )
        .join(models.LiquidacionVendedor,
              models.LiquidacionVendedor.id == models.LiquidacionContadoItem.liquidacion_id)
        .join(models.Talonera,
              models.Talonera.id == models.LiquidacionContadoItem.talonera_id)
        .join(models.Vendedor,
              models.Vendedor.id == models.LiquidacionVendedor.vendedor_id)
        .all()
    )

    if not lci_all:
        return JSONResponse({"ok": True, "items": [], "total": 0})

    asignados_slot1 = {
        (b.talonera_especial_id, b.numero_especial)
        for b in db.query(models.Boleta).filter(
            models.Boleta.comprador_id.isnot(None),
            models.Boleta.numero_especial.isnot(None),
        ).all()
    }
    asignados_slot2 = {
        (b.talonera_especial_2_id, b.numero_especial_2)
        for b in db.query(models.Boleta).filter(
            models.Boleta.comprador_id.isnot(None),
            models.Boleta.numero_especial_2.isnot(None),
        ).all()
    }

    items = []
    for lci, liq, tal, vend in lci_all:
        if (lci.talonera_id, lci.numero) in asignados_slot1:
            continue
        if (lci.talonera_id, lci.numero) in asignados_slot2:
            continue
        nd = tal.num_digitos or 3
        items.append({
            "talonera": tal.nombre,
            "numero": lci.numero,
            "numero_fmt": str(lci.numero).zfill(nd),
            "vendedor": vend.nombre,
            "fecha_liquidacion": liq.fecha.strftime("%d/%m/%Y") if liq.fecha else None,
            "liquidacion_id": liq.id,
        })

    # Ordenar por vendedor, luego por talonera y número
    items.sort(key=lambda x: (x["vendedor"], x["talonera"], x["numero"]))

    return JSONResponse({"ok": True, "items": items, "total": len(items)})


@router.post("/completar-vendedores")
async def completar_vendedores(request: Request, db: Session = Depends(get_db)):
    """
    Para cada comprador sin vendedor asignado:
    1. Si la zona tiene vendedor_id → usa ese.
    2. Si no → busca el vendedor más frecuente entre los otros socios de la misma zona.
    3. Asigna el vendedor a la boleta y, si la zona no tenía vendedor, lo vincula también.
    """
    from collections import Counter
    await auth_module.require_admin(request, db)

    # Todos los compradores con boleta pero sin vendedor y con zona
    sin_vendedor = (
        db.query(models.Comprador)
        .join(models.Boleta, models.Boleta.comprador_id == models.Comprador.id)
        .filter(models.Boleta.vendedor_id == None, models.Comprador.zona_id != None)
        .distinct()
        .all()
    )

    arreglados = 0
    for c in sin_vendedor:
        zona = db.query(models.Zona).get(c.zona_id)
        if not zona:
            continue

        vendedor_id = zona.vendedor_id  # primero el de la zona

        if not vendedor_id:
            # Buscar el vendedor más frecuente en la zona (excluyendo a este comprador)
            conteos = (
                db.query(models.Boleta.vendedor_id, func_count(models.Boleta.id))
                .join(models.Comprador, models.Comprador.id == models.Boleta.comprador_id)
                .filter(
                    models.Comprador.zona_id == zona.id,
                    models.Comprador.id != c.id,
                    models.Boleta.vendedor_id != None
                )
                .group_by(models.Boleta.vendedor_id)
                .order_by(func_count(models.Boleta.id).desc())
                .first()
            )
            if conteos:
                vendedor_id = conteos[0]
                # Vincular zona con este vendedor para el futuro
                zona.vendedor_id = vendedor_id

        if vendedor_id:
            for b in c.boletas:
                if not b.vendedor_id:
                    b.vendedor_id = vendedor_id
            arreglados += 1

    db.commit()
    from fastapi.responses import JSONResponse
    return JSONResponse({"ok": True, "arreglados": arreglados})


@router.post("/completar-cobradores")
async def completar_cobradores(request: Request, db: Session = Depends(get_db)):
    """
    Para cada comprador sin cobrador asignado:
    Usa el cobrador_id de la zona directamente.
    """
    await auth_module.require_admin(request, db)

    sin_cobrador = (
        db.query(models.Comprador)
        .join(models.Boleta, models.Boleta.comprador_id == models.Comprador.id)
        .filter(
            models.Boleta.cobrador_id == None,
            models.Comprador.zona_id != None,
            # Excluir boletas ya totalmente pagas (no necesitan cobrador)
            models.Boleta.cuotas_pagadas < models.Boleta.cuotas_pactadas,
        )
        .distinct()
        .all()
    )

    arreglados = 0
    for c in sin_cobrador:
        zona = db.query(models.Zona).get(c.zona_id)
        if not zona or not zona.cobrador_id:
            continue
        for b in c.boletas:
            # Solo asignar cobrador a boletas con cuotas pendientes
            if not b.cobrador_id and (b.cuotas_pagadas or 0) < (b.cuotas_pactadas or 0):
                b.cobrador_id = zona.cobrador_id
        arreglados += 1

    db.commit()
    from fastapi.responses import JSONResponse
    return JSONResponse({"ok": True, "arreglados": arreglados})


@router.post("/asignar-cobrador")
async def asignar_cobrador(request: Request, db: Session = Depends(get_db)):
    """
    Asigna (o transfiere) un cobrador a uno o varios compradores.

    Comportamiento (decidido con Sergio 15/05/2026):
    - Para cada boleta del comprador NO dada de baja:
        * setea b.cobrador_id = nuevo_cobrador_id
        * si la boleta ya estaba en una planilla y esa planilla NO está liquidada
          (no tiene Liquidacion asociada), la saca de esa planilla (planilla_id=None)
          para que el emplantillado del nuevo cobrador la incluya en el mes actual.
        * si la planilla YA está liquidada (mes cerrado), se respeta el historial y
          solo se cambia el cobrador para los meses futuros.
    - Las cuotas ya cobradas (historial_cuotas) se preservan.

    Body (form o JSON):
        comprador_ids: list[int]   (puede venir como múltiples 'comprador_ids' en form)
        cobrador_id:   int | "" (vacío = quitar cobrador)
    """
    from fastapi.responses import JSONResponse

    _user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')

    form = await request.form()
    ids_raw = form.getlist("comprador_ids") if hasattr(form, "getlist") else []
    if not ids_raw:
        # fallback: un solo id
        single = form.get("comprador_id")
        if single:
            ids_raw = [single]
    try:
        comprador_ids = [int(x) for x in ids_raw if str(x).strip()]
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "error": "IDs inválidos"}, status_code=400)

    if not comprador_ids:
        return JSONResponse({"ok": False, "error": "No se recibieron compradores"}, status_code=400)

    cobrador_id_raw = (form.get("cobrador_id") or "").strip()
    nuevo_cobrador_id: Optional[int] = None
    if cobrador_id_raw:
        try:
            nuevo_cobrador_id = int(cobrador_id_raw)
        except ValueError:
            return JSONResponse({"ok": False, "error": "cobrador_id inválido"}, status_code=400)

    # Validar que el cobrador exista y esté activo (si vino uno)
    cobrador_nombre = None
    if nuevo_cobrador_id is not None:
        cob = db.query(models.Cobrador).get(nuevo_cobrador_id)
        if not cob:
            return JSONResponse({"ok": False, "error": "Cobrador no encontrado"}, status_code=404)
        cobrador_nombre = cob.nombre

    # IDs de planillas que YA están liquidadas → no las tocamos
    planillas_liquidadas = {
        pid for (pid,) in db.query(models.Liquidacion.planilla_id)
        .filter(models.Liquidacion.planilla_id.isnot(None)).all()
    }

    asignados = 0       # boletas que recibieron un cobrador (nuevo o cambio)
    transferidos = 0    # boletas que se sacaron de una planilla abierta de otro cobrador
    sin_cambios = 0     # boletas que ya tenían ese cobrador
    socios_tocados = 0

    for cid in comprador_ids:
        c = db.query(models.Comprador).get(cid)
        if not c:
            continue
        toco_algo = False
        for b in c.boletas:
            # Respetar boletas dadas de baja (sticky, requieren acción explícita)
            if b.condicion == CondicionBoleta.BAJA:
                continue

            prev_cob = b.cobrador_id
            prev_planilla = b.planilla_id

            if prev_cob == nuevo_cobrador_id:
                sin_cambios += 1
                continue

            # Cambio de cobrador
            b.cobrador_id = nuevo_cobrador_id
            asignados += 1
            toco_algo = True

            # Si estaba en una planilla y esa planilla NO está cerrada → sacarla
            # (la planilla del nuevo cobrador la captará en el próximo emplantillado).
            if prev_planilla and prev_planilla not in planillas_liquidadas:
                b.planilla_id = None
                transferidos += 1

        if toco_algo:
            socios_tocados += 1

    db.commit()

    return JSONResponse({
        "ok": True,
        "socios_tocados": socios_tocados,
        "boletas_asignadas": asignados,
        "boletas_transferidas": transferidos,
        "sin_cambios": sin_cambios,
        "cobrador_nombre": cobrador_nombre,
    })


def _resolver_zona(zona_id: Optional[int], zona_nueva: str, db: Session) -> Optional[int]:
    """Devuelve el zona_id a usar: crea la zona si es nueva."""
    z_id, _ = _resolver_zona_ex(zona_id, zona_nueva, db)
    return z_id


def _resolver_zona_ex(zona_id: Optional[int], zona_nueva: str, db: Session):
    """Igual que _resolver_zona pero ademas devuelve el objeto Zona si fue
    recien creada (para que el endpoint pueda informarselo al frontend).
    Devuelve (zona_id, zona_creada_or_None).
    """
    nombre = zona_nueva.strip().upper() if zona_nueva else ""
    if nombre:
        z = db.query(models.Zona).filter(models.Zona.nombre == nombre).first()
        creada = None
        if not z:
            z = models.Zona(nombre=nombre)
            db.add(z)
            db.flush()
            creada = z
        return z.id, creada
    return (zona_id or None), None


@router.post("/crear")
async def crear(
    request: Request,
    apellido_nombre: str = Form(...),
    direccion: str = Form(""),
    zona_id: Optional[str] = Form(None),
    zona_nueva: str = Form(""),
    telefono: str = Form(""),
    fecha_compra: Optional[str] = Form(None),
    boleta_id: Optional[int] = Form(None),
    vendedor_id: Optional[int] = Form(None),
    cuotas_pactadas: Optional[int] = Form(None),
    cuotas_anticipadas: Optional[int] = Form(None),
    te1: Optional[str] = Form(None),   # talonera_especial_id   (CONTADO)
    ne1: Optional[str] = Form(None),   # numero_especial         (CONTADO)
    te2: Optional[str] = Form(None),   # talonera_especial_2_id (CONTADO 2 VECES)
    ne2: Optional[str] = Form(None),   # numero_especial_2       (CONTADO 2 VECES)
    db: Session = Depends(get_db)
):
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')
    zona, zona_creada = _resolver_zona_ex(_parse_zona_id(zona_id), zona_nueva, db)

    # ── Vínculo zona ↔ vendedor (una zona tiene un solo vendedor) ─────────
    effective_vendedor_id = vendedor_id
    if zona:
        z_obj = db.query(models.Zona).get(zona)
        if z_obj:
            if z_obj.vendedor_id:
                # La zona ya tiene vendedor asignado → usarlo siempre
                effective_vendedor_id = z_obj.vendedor_id
            elif vendedor_id:
                # La zona no tiene vendedor → vincularlo ahora
                z_obj.vendedor_id = vendedor_id
                effective_vendedor_id = vendedor_id

    c = models.Comprador(
        apellido_nombre=apellido_nombre.strip().upper(),
        direccion=direccion.strip().upper() or None,
        zona_id=zona,
        telefono=telefono.strip() or None
    )
    db.add(c)
    db.flush()
    if boleta_id:
        b = db.query(models.Boleta).get(boleta_id)
        if b:
            b.comprador_id = c.id
            b.fecha_venta = date.fromisoformat(fecha_compra) if fecha_compra else b.fecha_venta
            if effective_vendedor_id:
                b.vendedor_id = effective_vendedor_id
            # Si la boleta fue liquidada al contado (1 o 2 pagos) por el vendedor,
            # las cuotas YA están cobradas: forzar pactadas completas y todo pago
            # (ej: 12/12), sin importar lo que venga del formulario.
            # Se detecta el contado por DOS señales (cualquiera alcanza), porque no
            # siempre están las dos a la vez:
            #   - modalidad_liquidacion ('contado'/'contado2'): liquidaciones hechas
            #     desde la app (puede no tener numero_especial todavía).
            #   - numero_especial / numero_especial_2 asignado: contados de data vieja
            #     o rearmada por script, que NO setearon modalidad_liquidacion.
            _mod_liq = (b.modalidad_liquidacion or "").strip().lower()
            # Si el vendedor la liquidó EXPLÍCITAMENTE en cuotas, eso manda: NO se
            # fuerza a contado aunque haya quedado un número especial colgado por
            # error (dato viejo, o liquidación que después se corrigió a cuotas).
            # Se limpia ese número espurio para que la boleta quede coherente como
            # "en cuotas" y no se la marque "12/12 al contado". Antes alcanzaba con
            # un numero_especial para forzar el contado, lo que dejaba boletas en
            # cuotas atascadas como contado/emplanillables-no.
            if _mod_liq == "cuotas":
                if b.numero_especial is not None or b.numero_especial_2 is not None:
                    b.numero_especial = None
                    b.talonera_especial_id = None
                    b.numero_especial_2 = None
                    b.talonera_especial_2_id = None
                _contado_liq = False
            else:
                _contado_liq = (
                    _mod_liq in ("contado", "contado2")
                    or b.numero_especial is not None
                    or b.numero_especial_2 is not None
                )
            if _contado_liq:
                _nc_tal = (b.talonera.num_cuotas if b.talonera and b.talonera.num_cuotas else 0) \
                          or (cuotas_pactadas if cuotas_pactadas and cuotas_pactadas > 0 else 0) or 12
                b.cuotas_pactadas = _nc_tal
                ant = _nc_tal
            else:
                if cuotas_pactadas is not None and cuotas_pactadas > 0:
                    b.cuotas_pactadas = cuotas_pactadas
                ant = cuotas_anticipadas if cuotas_anticipadas and cuotas_anticipadas > 0 else 1
            b.cuotas_anticipadas = ant
            b.cuotas_pagadas = ant   # las cuotas anticipadas ya están cobradas

            # ── Números de sorteo especial (CONTADO / CONTADO 2 VECES) ──────
            # Vienen del modal Nuevo Socio cuando la boleta es al contado. Son
            # opcionales (los números llegan después, en talonarios de 5). Solo
            # se asignan si el número está libre (no usado por otra boleta).
            def _pi(x):
                try:
                    v = int(x)
                    return v if v > 0 else None
                except (ValueError, TypeError):
                    return None

            def _esp_libre(num, tal_id):
                if not num or not tal_id:
                    return False
                q = db.query(models.Boleta).filter(
                    models.Boleta.id != b.id,
                    (
                        ((models.Boleta.talonera_especial_id == tal_id) &
                         (models.Boleta.numero_especial == num)) |
                        ((models.Boleta.talonera_especial_2_id == tal_id) &
                         (models.Boleta.numero_especial_2 == num))
                    ),
                ).first()
                return q is None

            _te1, _ne1 = _pi(te1), _pi(ne1)
            _te2, _ne2 = _pi(te2), _pi(ne2)
            if _te1 and _ne1 and _esp_libre(_ne1, _te1):
                b.numero_especial = _ne1
                b.talonera_especial_id = _te1
            if _te2 and _ne2 and _esp_libre(_ne2, _te2):
                b.numero_especial_2 = _ne2
                b.talonera_especial_2_id = _te2

            # Auto-asignar cobrador según la zona del comprador,
            # SOLO si la boleta tiene cuotas pendientes para cobrar.
            # Las "al contado" (cuotas_pagadas >= cuotas_pactadas) NO necesitan
            # cobrador, aunque todavía no tengan numero_especial asignado
            # (se asigna recién al cerrarse el talonario especial).
            _toda_paga = (b.cuotas_pagadas or 0) >= (b.cuotas_pactadas or 0)
            if c.zona_id and not _toda_paga:
                z = db.query(models.Zona).get(c.zona_id)
                if z and z.cobrador_id:
                    b.cobrador_id = z.cobrador_id
            # Transición de condición según el flujo correcto:
            #   EN_COBRANZA = solo cuotas pendientes con cobrador asignado y no contado.
            #   VENDIDO     = todo lo demás (contado, cuotas finalizadas, sin cobrador).
            # Si la boleta venía de BAJA no se toca.
            _es_contado = (b.numero_especial is not None) or (b.numero_especial_2 is not None)
            _cuotas_pendientes = (b.cuotas_pagadas or 0) < (b.cuotas_pactadas or 0)
            _en_cobranza = bool(
                b.cobrador_id and _cuotas_pendientes and not _es_contado
            )
            if b.condicion in (
                CondicionBoleta.SIN_VENDER,
                CondicionBoleta.CAJA,
                CondicionBoleta.VENDIDO,
                CondicionBoleta.EN_COBRANZA,
            ):
                b.condicion = (
                    CondicionBoleta.EN_COBRANZA if _en_cobranza else CondicionBoleta.VENDIDO
                )
    db.commit()

    # Si el cliente espera JSON (modal AJAX), devolvemos info de la zona creada
    # para que el frontend pueda agregarla al dropdown sin recargar la pagina.
    accept = (request.headers.get("accept") or "").lower()
    quiere_json = (
        "application/json" in accept
        or request.headers.get("x-requested-with", "").lower() == "xmlhttprequest"
    )
    if quiere_json:
        nueva_zona_payload = None
        if zona_creada is not None:
            nueva_zona_payload = {
                "id": zona_creada.id,
                "nombre": zona_creada.nombre,
                "vendedor_id": zona_creada.vendedor_id,
            }
        return JSONResponse({
            "ok": True,
            "comprador_id": c.id,
            "zona_id": zona,
            "nueva_zona": nueva_zona_payload,
        })
    return RedirectResponse("/compradores/", status_code=302)


@router.post("/{comprador_id}/editar-basico")
async def editar_basico(
    comprador_id: int,
    request: Request,
    apellido_nombre: str = Form(...),
    direccion: str = Form(""),
    telefono: str = Form(""),
    zona_id: Optional[str] = Form(None),
    zona_nueva: str = Form(""),
    fecha_compra: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Edición RÁPIDA de un socio desde el modal de alta (flechas ◀ ▶).
    Solo actualiza datos básicos: apellido/nombre, dirección, teléfono, zona y la
    fecha de compra de sus boletas. NO toca vendedor, cobrador, condición, cuotas ni
    los números de boleta — para eso está el editor completo."""
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')
    c = db.query(models.Comprador).get(comprador_id)
    if not c:
        raise HTTPException(404, "Socio no encontrado")

    zona, zona_creada = _resolver_zona_ex(_parse_zona_id(zona_id), zona_nueva, db)
    c.apellido_nombre = apellido_nombre.strip().upper()
    c.direccion = direccion.strip().upper() or None
    c.telefono = telefono.strip() or None
    c.zona_id = zona
    if fecha_compra:
        try:
            _fv = date.fromisoformat(fecha_compra)
            for b in c.boletas:
                b.fecha_venta = _fv
        except ValueError:
            pass
    db.commit()

    nueva_zona_payload = None
    if zona_creada is not None:
        nueva_zona_payload = {
            "id": zona_creada.id,
            "nombre": zona_creada.nombre,
            "vendedor_id": zona_creada.vendedor_id,
        }
    return JSONResponse({
        "ok": True,
        "comprador_id": c.id,
        "zona_id": zona,
        "nueva_zona": nueva_zona_payload,
    })


def _derivar_condicion(boleta) -> str:
    """Calcula la condición de una boleta a partir de su estado real.

    Flujo de negocio (acordado con Sergio, mayo 2026):
      - SIN_VENDER : impresa, todavía no entregada a vendedor
      - CAJA       : en mano del vendedor, sin rendir ni vender
      - VENDIDO    : ya rendida a la institución; cubre tanto las que aún no tienen
                     socio como las de contado, y las de cuotas con todas las cuotas
                     pagas (categoría "contado / finalizada")
      - EN_COBRANZA: vendida en cuotas con cuotas pendientes y cobrador asignado,
                     y NO contado (sin numero_especial)
      - BAJA       : anulada (sticky)
    """
    if not boleta:
        return "SIN_VENDER"
    # BAJA es sticky.
    if boleta.condicion and boleta.condicion.value == "BAJA":
        return "BAJA"

    es_contado = (boleta.numero_especial is not None) or (boleta.numero_especial_2 is not None)
    cuotas_pendientes = (boleta.cuotas_pagadas or 0) < (boleta.cuotas_pactadas or 0)
    rendida = (boleta.liquidacion_vendedor_id is not None) or (boleta.comprador_id is not None)

    # Si está vendida y tiene cobrador con cuotas pendientes y NO es contado → EN_COBRANZA
    if rendida and boleta.cobrador_id and cuotas_pendientes and not es_contado:
        return "EN_COBRANZA"

    # Si está rendida (con o sin socio) y no es EN_COBRANZA → VENDIDO
    if rendida:
        return "VENDIDO"

    # No rendida: si la boleta tiene vendedor_id (= está en caja de algún vendedor) → CAJA
    if boleta.vendedor_id:
        return "CAJA"

    # Fallback
    return "SIN_VENDER"


@router.get("/{comprador_id}/editar", response_class=HTMLResponse)
async def editar_form(comprador_id: int, request: Request, db: Session = Depends(get_db)):
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')
    c = db.query(models.Comprador).get(comprador_id)
    if not c:
        raise HTTPException(404)
    zonas = db.query(models.Zona).order_by(models.Zona.nombre).all()
    vendedores = db.query(models.Vendedor).filter(models.Vendedor.activo == True).order_by(models.Vendedor.nombre).all()
    cobradores = db.query(models.Cobrador).filter(models.Cobrador.activo == True).order_by(models.Cobrador.nombre).all()
    # Taloneras COMUNES para reasignacion de numero (boton "asignar nueva talonera")
    taloneras_comunes = (
        db.query(models.Talonera)
        .filter(models.Talonera.tipo == "COMUN")
        .order_by(models.Talonera.multiplicador, models.Talonera.nombre)
        .all()
    )
    cond_derivada = _derivar_condicion(c.boletas[0]) if c.boletas else "SIN_VENDER"
    return templates.TemplateResponse(request, "comprador_editar.html", {
        "user": user,
        "comprador": c,
        "zonas": zonas,
        "vendedores": vendedores,
        "cobradores": cobradores,
        "taloneras_comunes": taloneras_comunes,
        "cond_derivada": cond_derivada,
    })


@router.get("/{comprador_id}/editar-fragmento", response_class=HTMLResponse)
async def editar_fragmento(comprador_id: int, request: Request, db: Session = Depends(get_db)):
    """Solo el formulario de edición (sin base.html), para inyectar en el modal
    nativo de la lista de socios."""
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')
    c = db.query(models.Comprador).get(comprador_id)
    if not c:
        raise HTTPException(404)
    zonas = db.query(models.Zona).order_by(models.Zona.nombre).all()
    vendedores = db.query(models.Vendedor).filter(models.Vendedor.activo == True).order_by(models.Vendedor.nombre).all()
    cobradores = db.query(models.Cobrador).filter(models.Cobrador.activo == True).order_by(models.Cobrador.nombre).all()
    taloneras_comunes = (
        db.query(models.Talonera)
        .filter(models.Talonera.tipo == "COMUN")
        .order_by(models.Talonera.multiplicador, models.Talonera.nombre)
        .all()
    )
    cond_derivada = _derivar_condicion(c.boletas[0]) if c.boletas else "SIN_VENDER"
    return templates.TemplateResponse(request, "_comprador_editar_form.html", {
        "comprador": c,
        "zonas": zonas,
        "vendedores": vendedores,
        "cobradores": cobradores,
        "taloneras_comunes": taloneras_comunes,
        "cond_derivada": cond_derivada,
    })


@router.post("/{comprador_id}/editar")
async def editar(
    comprador_id: int, request: Request,
    apellido_nombre: str = Form(...),
    direccion: str = Form(""),
    zona_id: Optional[str] = Form(None),
    zona_nueva: str = Form(""),
    telefono: str = Form(""),
    fecha_compra: Optional[str] = Form(None),
    boleta_id: Optional[int] = Form(None),
    vendedor_id: Optional[int] = Form(None),
    cobrador_id: Optional[int] = Form(None),
    cuotas_anticipadas: Optional[int] = Form(None),
    condicion: Optional[str] = Form(None),
    from_page: Optional[str] = Form(None),
    next: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')
    c = db.query(models.Comprador).get(comprador_id)
    if not c:
        raise HTTPException(404)
    zona = _resolver_zona(_parse_zona_id(zona_id), zona_nueva, db)
    c.apellido_nombre = apellido_nombre.strip().upper()
    c.direccion = direccion.strip().upper() or None
    c.zona_id = zona
    c.telefono = telefono.strip() or None

    # Si el usuario marca BAJA explícitamente, queda BAJA (sticky).
    # Si no, la condición se re-deriva al final basada en el estado real.
    marcar_baja = (condicion == "BAJA")
    reactivar = (condicion == "REACTIVAR")
    if marcar_baja:
        for b_exist in c.boletas:
            b_exist.condicion = CondicionBoleta.BAJA
    elif reactivar:
        # Limpiar BAJA — se re-derivará al final
        for b_exist in c.boletas:
            if b_exist.condicion == CondicionBoleta.BAJA:
                b_exist.condicion = CondicionBoleta.SIN_VENDER

    # Actualizar vendedor en todas las boletas existentes
    if vendedor_id:
        for b_exist in c.boletas:
            b_exist.vendedor_id = vendedor_id

    # Cobrador: si el usuario eligió uno manualmente, tiene prioridad;
    # si no, auto-asignar según la zona
    effective_cobrador_id = cobrador_id
    if not effective_cobrador_id and zona:
        z = db.query(models.Zona).get(zona)
        if z and z.cobrador_id:
            effective_cobrador_id = z.cobrador_id

    if effective_cobrador_id:
        for b_exist in c.boletas:
            b_exist.cobrador_id = effective_cobrador_id

    # Condición en boletas: la recalculamos al final del request (después de actualizar
    # cuotas_pagadas y la modalidad contado) en el helper de abajo. Ver bloque marcado
    # "RECALCULO FINAL DE CONDICION".

    # Actualizar cuotas_pagadas por boleta (campos cpag_<id>)
    form_data = await request.form()
    for b_exist in c.boletas:
        key = f"cpag_{b_exist.id}"
        if key in form_data:
            try:
                nuevas = int(form_data[key])
                if 0 <= nuevas <= (b_exist.cuotas_pactadas or 99):
                    b_exist.cuotas_pagadas = nuevas
                    # Mantener consistencia con la planilla/liquidación: las X
                    # negras de la planilla salen de cuotas_anticipadas, y las
                    # celdas con mes salen de historial_cuotas. Si acá se cargan
                    # 2 cuotas pagas (el vendedor cobró cuota 1 y 2 en la venta)
                    # y no se sincroniza anticipadas, la planilla dibuja UNA sola
                    # X y la próxima liquidación recalcula cuotas_pagadas hacia
                    # abajo, perdiendo la cuota. Ver mismo criterio en
                    # POST /{cid}/boleta/{bid}/cuotas.
                    import json as _json_cpag
                    _hl = 0
                    if b_exist.historial_cuotas:
                        try:
                            _hl = len(_json_cpag.loads(b_exist.historial_cuotas))
                        except (ValueError, TypeError):
                            _hl = 0
                    _ant_ok = max(1, nuevas - _hl)
                    if (b_exist.cuotas_anticipadas or 1) != _ant_ok:
                        b_exist.cuotas_anticipadas = _ant_ok
            except (ValueError, TypeError):
                pass

    # ── Modalidad CONTADO (ETAPA 2) ────────────────────────────────────────
    # Por boleta:
    #   modalidad_<id> = "cuotas" | "1pago" | "2pagos"
    #   te_<id>  / ne_<id>   = talonera_especial_id  / numero_especial   (CONTADO)
    #   te2_<id> / ne2_<id>  = talonera_especial_2_id / numero_especial_2 (CONTADO 2 VECES)
    # Reglas:
    #   - cuotas: limpia ambos slots
    #   - 1pago : asigna slot1 (CONTADO) y slot2 (CONTADO 2 VECES)
    #   - 2pagos: asigna solo slot2 (CONTADO 2 VECES); slot1 queda vacio
    # Validacion: si el numero ya esta asignado a OTRA boleta para esa talonera,
    # se rechaza (silencioso) y no se persiste.
    def _to_int(x):
        try:
            v = int(x)
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    def _esta_libre(num: int, talonera_id: int, except_boleta_id: int) -> bool:
        """¿Está libre este número (no asignado a otra boleta) para esa talonera?"""
        if not num or not talonera_id:
            return False
        q = db.query(models.Boleta).filter(
            models.Boleta.id != except_boleta_id,
            (
                ((models.Boleta.talonera_especial_id == talonera_id) &
                 (models.Boleta.numero_especial == num)) |
                ((models.Boleta.talonera_especial_2_id == talonera_id) &
                 (models.Boleta.numero_especial_2 == num))
            ),
        ).first()
        return q is None

    for b_exist in c.boletas:
        modal = (form_data.get(f"modalidad_{b_exist.id}") or "").strip().lower()
        if modal not in ("cuotas", "1pago", "2pagos"):
            continue  # no se mando, no tocamos
        te1 = _to_int(form_data.get(f"te_{b_exist.id}"))
        ne1 = _to_int(form_data.get(f"ne_{b_exist.id}"))
        te2 = _to_int(form_data.get(f"te2_{b_exist.id}"))
        ne2 = _to_int(form_data.get(f"ne2_{b_exist.id}"))

        if modal == "cuotas":
            b_exist.numero_especial = None
            b_exist.talonera_especial_id = None
            b_exist.numero_especial_2 = None
            b_exist.talonera_especial_2_id = None
            # CRÍTICO: limpiar también modalidad_liquidacion. Si queda 'contado'/
            # 'contado2', la boleta se vuelve a detectar como CONTADO (el modal la
            # re-muestra como "1 sólo pago" y, al recargar/cargar el socio, la
            # lógica de _contado_liq la fuerza de nuevo a 12/12 "al contado").
            # Por eso "se volvía a poner al contado" después de guardar.
            b_exist.modalidad_liquidacion = "cuotas"
            # Si venía forzada como contado (cuotas_anticipadas >= pactadas, ej 12/12),
            # el badge "CONTADO" se dispara por anticipo_total aunque ya no haya
            # número especial. Al volver a "en cuotas" reseteamos las anticipadas a 1
            # (cuota de la venta). Las cuotas pagadas reales se ajustan con el campo
            # "Cuotas pagas".
            if ((b_exist.cuotas_anticipadas or 0) >= (b_exist.cuotas_pactadas or 0)
                    and (b_exist.cuotas_pactadas or 0) > 0):
                b_exist.cuotas_anticipadas = 1
                if (b_exist.cuotas_pagadas or 0) >= (b_exist.cuotas_pactadas or 0):
                    b_exist.cuotas_pagadas = 1
        elif modal == "1pago":
            # slot 1
            if te1 and ne1 and _esta_libre(ne1, te1, b_exist.id):
                b_exist.numero_especial = ne1
                b_exist.talonera_especial_id = te1
            # slot 2
            if te2 and ne2 and _esta_libre(ne2, te2, b_exist.id):
                b_exist.numero_especial_2 = ne2
                b_exist.talonera_especial_2_id = te2
            b_exist.modalidad_liquidacion = "contado"
        elif modal == "2pagos":
            # solo slot 2
            b_exist.numero_especial = None
            b_exist.talonera_especial_id = None
            if te2 and ne2 and _esta_libre(ne2, te2, b_exist.id):
                b_exist.numero_especial_2 = ne2
                b_exist.talonera_especial_2_id = te2
            b_exist.modalidad_liquidacion = "contado2"

    # Agregar nueva boleta si se buscó una
    if boleta_id:
        b = db.query(models.Boleta).get(boleta_id)
        if b:
            b.comprador_id = comprador_id
            b.fecha_venta = date.fromisoformat(fecha_compra) if fecha_compra else b.fecha_venta
            if vendedor_id:
                b.vendedor_id = vendedor_id
            if cuotas_anticipadas and cuotas_anticipadas > 0:
                b.cuotas_anticipadas = cuotas_anticipadas
                b.cuotas_pagadas = cuotas_anticipadas
            if effective_cobrador_id:
                b.cobrador_id = effective_cobrador_id

    # Re-derivar condición al final (salvo BAJA explícita que ya quedó seteada)
    if not marcar_baja:
        for b_exist in c.boletas:
            if b_exist.condicion != CondicionBoleta.BAJA:
                derivada = _derivar_condicion(b_exist)
                try:
                    b_exist.condicion = CondicionBoleta(derivada)
                except ValueError:
                    pass

    db.commit()
    if next and next.startswith("/compradores"):
        redirect_url = next
    elif from_page == "mapa":
        redirect_url = "/compradores/mapa"
    else:
        redirect_url = "/compradores/"
    return RedirectResponse(redirect_url, status_code=302)


@router.post("/{comprador_id}/boleta/{boleta_id}/cuotas")
async def actualizar_cuotas(
    comprador_id: int, boleta_id: int, request: Request,
    cuotas_pagadas: int = Form(...),
    db: Session = Depends(get_db)
):
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')
    b = db.query(models.Boleta).filter(models.Boleta.id == boleta_id).first()
    if not b:
        from fastapi.responses import JSONResponse
        return JSONResponse({"ok": False, "error": f"boleta {boleta_id} no encontrada"}, status_code=404)
    b.cuotas_pagadas = cuotas_pagadas
    # Mantener consistencia con la planilla y la liquidación de cobranza:
    # las marcas de la planilla salen de cuotas_anticipadas (X negra) +
    # historial_cuotas (mes), y al liquidar una planilla se recalcula
    # cuotas_pagadas = anticipadas + len(historial). Una carga manual
    # (cuota cobrada FUERA de la cobranza, ej. por la institución) debe
    # reflejarse en anticipadas para que se dibuje la X y no se pierda
    # en la próxima liquidación.
    import json as _json
    _hist_len = 0
    if b.historial_cuotas:
        try:
            _hist_len = len(_json.loads(b.historial_cuotas))
        except Exception:
            _hist_len = 0
    _ant_consistente = max(1, (cuotas_pagadas or 0) - _hist_len)
    if (b.cuotas_anticipadas or 1) != _ant_consistente:
        b.cuotas_anticipadas = _ant_consistente
    db.commit()
    from fastapi.responses import JSONResponse
    return JSONResponse({"ok": True, "cuotas_pagadas": b.cuotas_pagadas})


@router.post("/{comprador_id}/boleta/{boleta_id}/reasignar-talonera")
async def reasignar_talonera(
    comprador_id: int, boleta_id: int, request: Request,
    nueva_talonera_id: int = Form(...),
    nuevo_numero: int = Form(...),
    db: Session = Depends(get_db)
):
    """Reasigna una boleta del socio a OTRO numero/talonera.
    Caso de uso: el operador cargo el socio con el numero de talonera/boleta
    equivocado y necesita corregirlo despues. Decision (Sergio 16/05/2026):
    en lugar de mover datos entre dos registros (con sus referencias a
    LiquidacionDetalle, planillas, etc.) hacemos un INTERCAMBIO de
    (talonera_id, numero_principal, numeros_adicionales) entre la boleta
    actual del socio (B1) y la boleta libre destino (B2). Asi:
      - B1 conserva su id y todas las referencias externas, pero pasa a
        representar el numero CORRECTO.
      - B2 queda con la identidad del numero VIEJO (sigue SIN_VENDER, libre)
        y puede ser asignada a otro socio en el futuro.
    Validaciones:
      - La boleta destino DEBE existir como Boleta pre-generada.
      - Debe estar libre (sin comprador_id, condicion SIN_VENDER).
      - La talonera destino debe ser COMUN (no CONTADO).
      - El numero debe estar dentro del rango de la talonera.
    """
    from fastapi.responses import JSONResponse
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')

    b1 = db.query(models.Boleta).get(boleta_id)
    if not b1 or b1.comprador_id != comprador_id:
        return JSONResponse({"ok": False, "error": "Boleta no encontrada para este socio"}, status_code=404)

    nueva_t = db.query(models.Talonera).get(nueva_talonera_id)
    if not nueva_t:
        return JSONResponse({"ok": False, "error": "Talonera no encontrada"}, status_code=404)
    if (nueva_t.tipo or "COMUN") != "COMUN":
        return JSONResponse({"ok": False, "error": "Solo se puede reasignar a taloneras COMUNES"}, status_code=400)

    if nueva_t.numero_inicio is not None and nueva_t.numero_fin is not None:
        if not (nueva_t.numero_inicio <= nuevo_numero <= nueva_t.numero_fin):
            return JSONResponse({
                "ok": False,
                "error": f"El número {nuevo_numero} está fuera del rango de la talonera "
                         f"({nueva_t.numero_inicio}-{nueva_t.numero_fin})"
            }, status_code=400)

    b2 = db.query(models.Boleta).filter(
        models.Boleta.talonera_id == nueva_talonera_id,
        models.Boleta.numero_principal == nuevo_numero,
    ).first()

    if not b2:
        return JSONResponse({
            "ok": False,
            "error": f"El número {nuevo_numero} no existe en {nueva_t.nombre}. "
                     f"Primero generá las boletas en /taloneras/."
        }, status_code=404)

    if b2.id == b1.id:
        return JSONResponse({"ok": False, "error": "Es la misma boleta — no hay nada que cambiar"}, status_code=400)

    # No permitir intercambiar con una boleta que ya tiene socio asignado
    if b2.comprador_id is not None:
        comp_other = db.query(models.Comprador).get(b2.comprador_id)
        nombre_other = comp_other.apellido_nombre if comp_other else f"socio #{b2.comprador_id}"
        return JSONResponse({
            "ok": False,
            "error": f"El número {nuevo_numero} de {nueva_t.nombre} ya está asignado a: {nombre_other}"
        }, status_code=400)

    # Estado no-libre: SIN_VENDER es el unico estado valido para una boleta destino
    if b2.condicion and b2.condicion != CondicionBoleta.SIN_VENDER:
        return JSONResponse({
            "ok": False,
            "error": f"El número {nuevo_numero} de {nueva_t.nombre} no está libre "
                     f"(estado: {b2.condicion.value})"
        }, status_code=400)

    # ── INTERCAMBIO de identidad de talonera/numero entre B1 y B2 ──────────
    from .taloneras import calcular_numeros

    old_tal_id = b1.talonera_id
    old_num    = b1.numero_principal
    old_adic   = b1.numeros_adicionales

    # B2 toma la identidad VIEJA (libre, en el lugar viejo)
    b2.talonera_id        = old_tal_id
    b2.numero_principal   = old_num
    b2.numeros_adicionales = old_adic

    # B1 toma la identidad NUEVA y mantiene todos sus datos de socio
    b1.talonera_id      = nueva_talonera_id
    b1.numero_principal = nuevo_numero
    b1.numeros_adicionales = (
        calcular_numeros(nuevo_numero, nueva_t.num_series or 1, nueva_t.offset_series or 0) or None
    )

    db.commit()

    return JSONResponse({
        "ok": True,
        "boleta_id": b1.id,
        "nuevo_numero": nuevo_numero,
        "talonera_nombre": nueva_t.nombre,
        "numeros_adicionales": b1.numeros_adicionales or "",
    })


def _reset_boleta_a_stock(db: Session, b) -> None:
    """Devuelve una boleta al stock de la institucion, LIMPIA (fix A-3).

    Se usa al "liberar" una boleta y al ELIMINAR un socio. Antes, eliminar
    un socio solo borraba comprador_id/fecha/condicion y dejaba una boleta
    fantasma: seguia en su planilla (sumando en el consolidado), conservaba
    historial, numeros especiales y liquidacion_vendedor_id (inflando el
    "Total liquidados" del dashboard) y no se podia volver a entregar limpia.

    Pasos:
      1. Descuenta la boleta de su LiquidacionVendedor (snapshots + montos),
         mismo helper que usa "sacar numero" - evita el drift del dashboard.
      2. Resetea TODOS los campos de venta/cobranza al estado de fabrica.
    """
    if b.liquidacion_vendedor_id:
        _liq = db.query(models.LiquidacionVendedor).get(b.liquidacion_vendedor_id)
        if _liq:
            from .vendedores import _quitar_boleta_de_liq
            _quitar_boleta_de_liq(_liq, b)

    b.comprador_id          = None
    b.vendedor_id           = None
    b.cobrador_id           = None
    b.planilla_id           = None
    b.fecha_venta           = None
    b.condicion             = CondicionBoleta.SIN_VENDER
    b.cuotas_pagadas        = 0
    b.cuotas_anticipadas    = 1
    b.total_pagado          = 0.0
    b.historial_cuotas      = None
    b.mes_baja              = None
    b.numero_especial       = None
    b.talonera_especial_id  = None
    b.numero_especial_2     = None
    b.talonera_especial_2_id = None
    b.liquidacion_vendedor_id = None
    b.modalidad_liquidacion = None


@router.post("/{comprador_id}/boleta/{boleta_id}/liberar")
async def liberar_boleta(
    comprador_id: int, boleta_id: int, request: Request,
    confirmacion: str = Form(""),
    db: Session = Depends(get_db)
):
    """Libera una boleta: desvincula al socio, resetea TODOS los datos de venta
    y la devuelve al estado SIN_VENDER. La talonera vuelve a estar disponible
    para asignar a otro socio.

    Doble seguro:
      - El cliente debe enviar `confirmacion` con el numero_principal exacto
        de la boleta (ej: "8608"). Sin eso, no se libera.
      - El JS de la UI ya tiene un confirm() previo + input tipeado.

    Si el socio queda sin boletas tras la liberacion, tambien se elimina el
    registro del Comprador (mismo criterio que el endpoint /eliminar).

    Se borran:
      - comprador_id, vendedor_id, cobrador_id, planilla_id
      - fecha_venta, condicion (SIN_VENDER), historial_cuotas
      - cuotas_pagadas (0), cuotas_anticipadas (1), total_pagado (0)
      - numero_especial, talonera_especial_id, numero_especial_2, talonera_especial_2_id
      - liquidacion_vendedor_id

    Decision (Sergio 16/05/2026): este boton es para CORREGIR cargas erroneas,
    no para dar de baja por morosidad. Por eso resetea todo. Para baja por
    morosidad existe el boton "Dar de baja" que preserva el historial.
    """
    from fastapi.responses import JSONResponse
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')

    b = db.query(models.Boleta).get(boleta_id)
    if not b or b.comprador_id != comprador_id:
        return JSONResponse({"ok": False, "error": "Boleta no encontrada para este socio"}, status_code=404)

    # Doble seguro: confirmacion debe ser exactamente el numero_principal
    confirmacion_esperada = str(b.numero_principal)
    if (confirmacion or "").strip() != confirmacion_esperada:
        return JSONResponse({
            "ok": False,
            "error": f"Confirmación inválida. Debes tipear exactamente '{confirmacion_esperada}' para confirmar."
        }, status_code=400)

    c = db.query(models.Comprador).get(comprador_id)
    if not c:
        return JSONResponse({"ok": False, "error": "Socio no encontrado"}, status_code=404)

    talonera_nombre = b.talonera.nombre if b.talonera else ""
    numero_liberado = b.numero_principal

    # ── Descuento de liquidacion + reset COMPLETO (helper compartido) ──────
    _reset_boleta_a_stock(db, b)

    # ── Si el socio queda sin boletas, eliminarlo ─────────────────────────
    db.flush()
    socio_eliminado = False
    boletas_restantes = (
        db.query(models.Boleta)
        .filter(models.Boleta.comprador_id == comprador_id)
        .count()
    )
    if boletas_restantes == 0:
        db.delete(c)
        socio_eliminado = True

    db.commit()

    return JSONResponse({
        "ok": True,
        "numero_liberado": numero_liberado,
        "talonera_nombre": talonera_nombre,
        "socio_eliminado": socio_eliminado,
        "boletas_restantes": boletas_restantes,
    })


@router.post("/{comprador_id}/eliminar")
async def eliminar(comprador_id: int, request: Request, db: Session = Depends(get_db)):
    _perm_user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(_perm_user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')
    c = db.query(models.Comprador).get(comprador_id)
    if c:
        # Reset COMPLETO de cada boleta (fix A-3): antes quedaban "fantasma"
        # (en planilla, con historial y liquidacion colgando). Ver helper.
        for b in c.boletas:
            _reset_boleta_a_stock(db, b)
        db.delete(c)
        db.commit()
    return RedirectResponse("/compradores/", status_code=302)


from fastapi.responses import JSONResponse


@router.get("/exportar")
async def exportar_excel(request: Request, db: Session = Depends(get_db)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    await auth_module.require_user(request, db)

    compradores = (db.query(models.Comprador)
                   .join(models.Zona, isouter=True)
                   .order_by(models.Comprador.apellido_nombre)
                   .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Socios"

    # ── Estilos ──────────────────────────────────────────────────────────
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill    = PatternFill("solid", start_color="C00000")
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left",   vertical="center")
    thin           = Side(style="thin", color="CCCCCC")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font      = Font(name="Arial", size=9)
    alt_fill       = PatternFill("solid", start_color="FFF0F0")

    # ── Encabezados ──────────────────────────────────────────────────────
    headers = [
        ("N° Boleta",       12),
        ("Talonera",        10),
        ("Apellido y Nombre", 28),
        ("Dirección",       28),
        ("Zona",            14),
        ("Fecha Compra",    13),
        ("Vendedor",        18),
        ("Cobrador",        18),
        ("Teléfono",        13),
        ("Cuotas Pactadas",  8),
        ("Al Contado",       8),
        ("Condición",       13),
    ]

    for col_idx, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # ── Datos ─────────────────────────────────────────────────────────────
    row_num = 2
    for c in compradores:
        for b in (c.boletas if c.boletas else [None]):
            fill = alt_fill if row_num % 2 == 0 else None

            def cell(col, value, align=left):
                ce = ws.cell(row=row_num, column=col, value=value)
                ce.font      = data_font
                ce.alignment = align
                ce.border    = border
                if fill:
                    ce.fill = fill
                return ce

            cell(1,  b.numero_principal if b else "",                                center)
            cell(2,  b.talonera.nombre  if b and b.talonera  else "",                center)
            cell(3,  c.apellido_nombre)
            cell(4,  c.direccion or "")
            cell(5,  c.zona.nombre      if c.zona            else "")
            cell(6,  b.fecha_venta.strftime("%d/%m/%Y") if b and b.fecha_venta else "", center)
            cell(7,  b.vendedor.nombre  if b and b.vendedor  else "")
            cell(8,  b.cobrador.nombre  if b and b.cobrador  else "")
            cell(9,  c.telefono or "")
            cell(10, b.cuotas_pactadas   if b else "",                               center)
            cell(11, b.cuotas_anticipadas if b else "",                              center)
            cell(12, b.condicion.value   if b and b.condicion else "",               center)

            ws.row_dimensions[row_num].height = 15
            row_num += 1

    # ── Autofilter ───────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:L{row_num - 1}"

    # ── Stream ───────────────────────────────────────────────────────────
    try:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(500, detail=f"Error generando el archivo Excel: {exc}")

    from datetime import date as dt
    filename = f"socios_{dt.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/boleta/{boleta_id}/contado-disponibles")
async def contado_disponibles(boleta_id: int, request: Request, db: Session = Depends(get_db)):
    """Devuelve los numeros CONTADO y CONTADO 2 VECES que el vendedor de la
    boleta tiene en mano (entregados via EntregaCaja menos los ya asignados a
    cualquier boleta). Usado por la pantalla de editar comprador para que el
    operador elija manualmente que numero del pool del vendedor le corresponde
    a este socio segun la modalidad de pago.
    Respuesta:
      {
        "ok": true,
        "vendedor_id": int|null,
        "vendedor_nombre": str|null,
        "current": {"numero_especial": int|null, "talonera_especial_id": int|null,
                    "numero_especial_2": int|null, "talonera_especial_2_id": int|null},
        "taloneras_contado": [
          {"id": int, "nombre": str, "color": str, "num_digitos": int,
           "rol": "CONTADO"|"CONTADO_2"|"OTRO",
           "numeros_libres": [int, ...]},
          ...
        ]
      }
    rol se infiere por el nombre: si el nombre normalizado contiene "2" => CONTADO_2,
    si arranca con "CONTADO" sin numero => CONTADO, sino OTRO.
    """
    from fastapi.responses import JSONResponse
    await auth_module.require_user(request, db)
    b = db.query(models.Boleta).get(boleta_id)
    if not b:
        raise HTTPException(404, "Boleta no encontrada")

    vid = b.vendedor_id
    v = db.query(models.Vendedor).get(vid) if vid else None

    # Taloneras CONTADO existentes
    taloneras_c = db.query(models.Talonera).filter(models.Talonera.tipo == "CONTADO").all()

    def _rol(nombre: str) -> str:
        import re as _re
        nm = (nombre or "").strip().upper()
        if not nm.startswith("CONTADO"):
            return "OTRO"
        m = _re.search(r"(\d+)", nm)
        if m and int(m.group(1)) >= 2:
            return "CONTADO_2"
        return "CONTADO"

    # Entregas de este vendedor para taloneras CONTADO
    entregas_v = []
    if vid:
        entregas_v = db.query(models.EntregaCaja).filter_by(vendedor_id=vid).all()

    # Map nombre normalizado -> talonera para hacer match con entregas
    norm_map = {(t.nombre or "").strip().lower(): t for t in taloneras_c}

    # Numeros que el vendedor YA declaro como vendidos en una liquidacion
    # (LiquidacionContadoItem). Solo estos pueden aparecer para asignar al socio:
    # los entregados pero no liquidados todavia siguen en su caja y no se cargan.
    # Mapeo: talonera_id -> set(numeros liquidados por este vendedor).
    nums_liquidados_por_tal: dict = {}
    if vid:
        try:
            _rows_liq = db.query(
                models.LiquidacionContadoItem.talonera_id,
                models.LiquidacionContadoItem.numero,
            ).join(
                models.LiquidacionVendedor,
                models.LiquidacionVendedor.id == models.LiquidacionContadoItem.liquidacion_id,
            ).filter(
                models.LiquidacionVendedor.vendedor_id == vid,
            ).all()
            for tid, num in _rows_liq:
                nums_liquidados_por_tal.setdefault(int(tid), set()).add(int(num))
        except Exception:
            nums_liquidados_por_tal = {}

    # Para cada talonera CONTADO, computar los numeros entregados a este vendedor
    out_taloneras = []
    for t in taloneras_c:
        nombre_norm = (t.nombre or "").strip().lower()
        rangos = []
        for e in entregas_v:
            if (e.talonera_nombre or "").strip().lower() == nombre_norm:
                rangos.append((int(e.desde), int(e.hasta)))
        nums_entregados = set()
        for d, h in rangos:
            if h < d:
                continue
            nums_entregados.update(range(d, h + 1))
        if not nums_entregados:
            # El vendedor no tiene numeros de esta talonera; si la boleta YA
            # tiene asignado un numero de esta talonera lo incluimos igual.
            asignados_aqui = set()
            if b.talonera_especial_id == t.id and b.numero_especial:
                asignados_aqui.add(b.numero_especial)
            if b.talonera_especial_2_id == t.id and b.numero_especial_2:
                asignados_aqui.add(b.numero_especial_2)
            if not asignados_aqui:
                continue
            libres = sorted(asignados_aqui)
        else:
            # Numeros ya asignados a alguna boleta para esta talonera
            asignados_rows = db.query(
                models.Boleta.numero_especial,
                models.Boleta.talonera_especial_id,
                models.Boleta.numero_especial_2,
                models.Boleta.talonera_especial_2_id,
                models.Boleta.id,
            ).filter(
                ((models.Boleta.talonera_especial_id == t.id) &
                 models.Boleta.numero_especial.isnot(None)) |
                ((models.Boleta.talonera_especial_2_id == t.id) &
                 models.Boleta.numero_especial_2.isnot(None))
            ).all()
            asignados_otros = set()
            for ne, tei, ne2, tei2, bid in asignados_rows:
                if bid == b.id:
                    continue  # no contar lo que ya asigne a esta misma boleta
                if tei == t.id and ne is not None:
                    asignados_otros.add(int(ne))
                if tei2 == t.id and ne2 is not None:
                    asignados_otros.add(int(ne2))
            # Solo numeros LIQUIDADOS por el vendedor pueden aparecer en el dropdown.
            # Excepcion: el numero actual de la boleta se mantiene aunque no este
            # liquidado (compatibilidad con datos cargados antes de esta regla).
            nums_liq_t = nums_liquidados_por_tal.get(t.id, set())
            libres_base = (nums_entregados & nums_liq_t) - asignados_otros
            if b.talonera_especial_id == t.id and b.numero_especial:
                libres_base.add(int(b.numero_especial))
            if b.talonera_especial_2_id == t.id and b.numero_especial_2:
                libres_base.add(int(b.numero_especial_2))
            libres = sorted(libres_base)

        out_taloneras.append({
            "id": t.id,
            "nombre": t.nombre,
            "color": t.color or "#fff8e1",
            "num_digitos": t.num_digitos or 3,
            "rol": _rol(t.nombre),
            "numeros_libres": libres,
        })

    # Orden: CONTADO primero, CONTADO_2 segundo, OTRO al final
    out_taloneras.sort(key=lambda x: (
        0 if x["rol"] == "CONTADO" else 1 if x["rol"] == "CONTADO_2" else 2,
        x["nombre"],
    ))

    return JSONResponse({
        "ok": True,
        "vendedor_id": vid,
        "vendedor_nombre": v.nombre if v else None,
        "current": {
            "numero_especial": b.numero_especial,
            "talonera_especial_id": b.talonera_especial_id,
            "numero_especial_2": b.numero_especial_2,
            "talonera_especial_2_id": b.talonera_especial_2_id,
        },
        "taloneras_contado": out_taloneras,
    })


@router.get("/{comprador_id}/detalle")
async def detalle_json(comprador_id: int, request: Request, db: Session = Depends(get_db)):
    await auth_module.require_user(request, db)
    c = db.query(models.Comprador).get(comprador_id)
    if not c:
        raise HTTPException(404)

    boletas = []
    for b in c.boletas:
        numeros = [b.numero_principal]
        if b.numeros_adicionales:
            for n in b.numeros_adicionales.split(", "):
                try:
                    numeros.append(int(n.strip()))
                except ValueError:
                    pass
        boletas.append({
            "id": b.id,
            "talonera": b.talonera.nombre if b.talonera else None,
            "numero_principal": b.numero_principal,
            "numeros_adicionales": b.numeros_adicionales or "",
            "todos_numeros": numeros,
            "fecha_venta": b.fecha_venta.strftime("%d/%m/%Y") if b.fecha_venta else None,
            "condicion": b.condicion,
            "vendedor": b.vendedor.nombre if b.vendedor else None,
            "cobrador": b.cobrador.nombre if b.cobrador else None,
            "cuotas_pactadas": b.cuotas_pactadas,
            "cuotas_pagadas": b.cuotas_pagadas,
            "total_pagado": b.total_pagado,
        })

    return JSONResponse({
        "id": c.id,
        "apellido_nombre": c.apellido_nombre,
        "direccion": c.direccion or "",
        "telefono": c.telefono or "",
        "zona": c.zona.nombre if c.zona else None,
        "boletas": boletas,
    })


# ── MAPA DE SOCIOS ──────────────────────────────────────────────────────
# Vista de mapa con la ubicación de los números vendidos en Concepción del
# Uruguay y alrededores. El color del marcador se toma de Talonera.color
# (picker de la talonera) para distinguir las PATAs.

@router.get("/mapa", response_class=HTMLResponse)
async def mapa_view(request: Request, db: Session = Depends(get_db)):
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')
    return templates.TemplateResponse(request, "compradores_mapa.html", {
        "user": user,
    })


def _norm_direccion(s: str) -> str:
    """Normalizacion para usar como clave en geocode_cache."""
    return " ".join((s or "").strip().upper().split())


@router.get("/mapa-data")
async def mapa_data(request: Request, db: Session = Depends(get_db)):
    """Devuelve JSON con los puntos a graficar en el mapa.

    Una entrada por cada boleta vendida (comprador_id IS NOT NULL) cuya talonera
    sea COMUN y cuyo comprador tenga direccion cargada. El color del marcador
    proviene de Talonera.color (picker en UI); si la talonera no tiene color
    propio se cae a una paleta por PATA.

    Si la direccion ya fue geocodificada y guardada en GeocodeCache, se incluyen
    lat/lng pre-cargados para evitar que el cliente vuelva a consultar Nominatim.
    """
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')

    # Color = Talonera.color (picker de la UI de Taloneras). Si está vacío
    # o quedó en blanco por defecto, usamos un gris neutro como señal de
    # "esta talonera no tiene color asignado, andá a Taloneras a setearlo".
    SIN_COLOR_FALLBACK = "#9e9e9e"

    rows = (
        db.query(
            models.Boleta, models.Comprador, models.Talonera,
            models.Zona, models.Vendedor, models.Cobrador,
        )
        .join(models.Comprador, models.Boleta.comprador_id == models.Comprador.id)
        .join(models.Talonera, models.Boleta.talonera_id == models.Talonera.id)
        .outerjoin(models.Zona, models.Comprador.zona_id == models.Zona.id)
        .outerjoin(models.Vendedor, models.Boleta.vendedor_id == models.Vendedor.id)
        .outerjoin(models.Cobrador, models.Boleta.cobrador_id == models.Cobrador.id)
        .filter(models.Comprador.direccion.isnot(None))
        .filter(models.Comprador.direccion != "")
        .filter(models.Talonera.tipo == "COMUN")
        .order_by(models.Talonera.nombre, models.Boleta.numero_principal)
        .all()
    )

    # Pre-cargar todas las direcciones cacheadas en un solo SELECT (clave -> (lat,lng,intento))
    cache_rows = db.query(
        models.GeocodeCache.direccion,
        models.GeocodeCache.lat,
        models.GeocodeCache.lng,
    ).all()
    cache_map = {}
    for d, lat, lng in cache_rows:
        cache_map[d] = (lat, lng)

    puntos = []
    patas_set = {}
    vendedores_set = {}
    cobradores_set = {}
    for b, c, t, z, v, cob in rows:
        color = (t.color or "").strip().lower()
        if not color or color in ("#ffffff", "#fff", "white"):
            color = SIN_COLOR_FALLBACK

        # Formato de numero con cifras de la talonera (default 4 para COMUN)
        try:
            num_digitos = t.num_digitos or 4
        except Exception:
            num_digitos = 4
        numero_fmt = str(b.numero_principal).zfill(num_digitos)

        direccion_norm = _norm_direccion(c.direccion or "")
        cached = cache_map.get(direccion_norm)
        lat_pre = cached[0] if cached else None
        lng_pre = cached[1] if cached else None
        # cache_map devuelve (None, None) cuando la direccion ya se intento y fallo
        ya_intentado = direccion_norm in cache_map

        puntos.append({
            "boleta_id": b.id,
            "comprador_id": c.id,
            "numero": numero_fmt,
            "numero_int": b.numero_principal,
            "apellido_nombre": c.apellido_nombre,
            "direccion": c.direccion or "",
            "zona": z.nombre if z else None,
            "vendedor_id": v.id if v else None,
            "vendedor_nombre": v.nombre if v else None,
            "cobrador_id": cob.id if cob else None,
            "cobrador_nombre": cob.nombre if cob else None,
            "pata": t.nombre,
            "pata_label": (t.nombre or "").replace("PATA ", "X"),
            "color": color,
            "lat": lat_pre,
            "lng": lng_pre,
            "ya_intentado": ya_intentado,  # true = no reintentar geocoding (fallo previo)
            "es_contado": (b.numero_especial is not None) or ((b.cuotas_anticipadas or 0) >= (b.cuotas_pactadas or 1)),
            "contado_2_veces": b.numero_especial_2 is not None,
        })
        if t.nombre not in patas_set:
            patas_set[t.nombre] = color
        if v and v.id not in vendedores_set:
            vendedores_set[v.id] = v.nombre
        if cob and cob.id not in cobradores_set:
            cobradores_set[cob.id] = cob.nombre

    patas = [
        {"nombre": k, "label": k.replace("PATA ", "X"), "color": v}
        for k, v in sorted(patas_set.items())
    ]
    vendedores = [
        {"id": vid, "nombre": vnom}
        for vid, vnom in sorted(vendedores_set.items(), key=lambda x: (x[1] or "").upper())
    ]
    cobradores = [
        {"id": cid, "nombre": cnom}
        for cid, cnom in sorted(cobradores_set.items(), key=lambda x: (x[1] or "").upper())
    ]

    return JSONResponse({
        "ok": True,
        "centro": {"lat": -32.4847, "lng": -58.2347, "nombre": "Concepción del Uruguay"},
        "patas": patas,
        "vendedores": vendedores,
        "cobradores": cobradores,
        "puntos": puntos,
        "total_cacheadas": sum(1 for p in puntos if p["lat"] is not None),
        "total_pendientes": sum(
            1 for p in puntos if p["lat"] is None and not p["ya_intentado"]
        ),
    })


@router.post("/mapa-geocode-save")
async def mapa_geocode_save(request: Request, db: Session = Depends(get_db)):
    """Guarda en GeocodeCache las coords obtenidas por el cliente desde Nominatim.

    Body JSON: {"direccion": "...", "lat": -32.48, "lng": -58.23}
    Si lat/lng vienen en null, marca la direccion como "no ubicable" (no reintentar).
    """
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')

    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)

    direccion = _norm_direccion(payload.get("direccion") or "")
    if not direccion:
        return JSONResponse({"ok": False, "error": "direccion vacía"}, status_code=400)

    lat = payload.get("lat")
    lng = payload.get("lng")
    try:
        lat = float(lat) if lat is not None else None
        lng = float(lng) if lng is not None else None
    except (TypeError, ValueError):
        lat, lng = None, None

    # Bbox sanity check — Concepcion del Uruguay y alrededores
    if lat is not None and lng is not None:
        if not (-33.0 <= lat <= -32.0 and -58.7 <= lng <= -57.8):
            lat, lng = None, None  # fuera del area: tratar como fallo

    from datetime import datetime as _dt
    entry = db.query(models.GeocodeCache).filter(
        models.GeocodeCache.direccion == direccion
    ).first()
    if entry:
        entry.lat = lat
        entry.lng = lng
        entry.intentos = (entry.intentos or 0) + 1
        entry.last_try = _dt.now()
    else:
        entry = models.GeocodeCache(
            direccion=direccion, lat=lat, lng=lng, intentos=1, last_try=_dt.now(),
        )
        db.add(entry)
    db.commit()
    return JSONResponse({"ok": True, "direccion": direccion, "lat": lat, "lng": lng})


@router.post("/mapa-geocode-reset")
async def mapa_geocode_reset(request: Request, db: Session = Depends(get_db)):
    """Borra todo el cache server-side de geocoding. Requiere admin."""
    await auth_module.require_admin(request, db)
    n = db.query(models.GeocodeCache).delete()
    db.commit()
    return JSONResponse({"ok": True, "borradas": n})


@router.post("/mapa-geocode-retry")
async def mapa_geocode_retry(request: Request, db: Session = Depends(get_db)):
    """Borra del cache SOLO las direcciones que fallaron (lat IS NULL),
    para que la próxima carga del mapa las vuelva a intentar con Nominatim.
    Útil después de corregir direcciones en Socios.
    """
    await auth_module.require_user(request, db)
    n = db.query(models.GeocodeCache).filter(
        models.GeocodeCache.lat.is_(None)
    ).delete()
    db.commit()
    return JSONResponse({"ok": True, "borradas": n})


@router.post("/{cid}/actualizar-direccion-mapa")
async def actualizar_direccion_mapa(cid: int, request: Request, db: Session = Depends(get_db)):
    """Actualiza solo la dirección de un socio (usado desde el panel
    'No ubicadas' del mapa para edición inline) y borra del GeocodeCache
    la entrada de la dirección vieja, para que la nueva se geocodifique.
    """
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'editar'):
        raise HTTPException(403, 'No tenés permiso para editar en esta sección')

    # Aceptar tanto form como JSON
    try:
        payload = await request.json()
    except Exception:
        form = await request.form()
        payload = {"direccion": form.get("direccion") or ""}

    nueva = (payload.get("direccion") or "").strip().upper()
    if not nueva:
        return JSONResponse({"ok": False, "error": "Dirección vacía"}, status_code=400)

    c = db.query(models.Comprador).get(cid)
    if not c:
        return JSONResponse({"ok": False, "error": "Socio no encontrado"}, status_code=404)

    direccion_vieja_norm = _norm_direccion(c.direccion or "")
    c.direccion = nueva
    nueva_norm = _norm_direccion(nueva)

    # Reset del cache para esta edición:
    #   - Siempre borramos la entrada de la dirección vieja (ya no la usa nadie
    #     desde este socio; si otro socio la sigue usando se re-geocodificará
    #     en la próxima carga del mapa).
    #   - Siempre borramos la entrada de la dirección nueva (aunque estuviera
    #     ya cacheada con coords). Esto fuerza una regeocodificación fresca
    #     después de editar.
    borradas = 0
    if direccion_vieja_norm:
        borradas += db.query(models.GeocodeCache).filter(
            models.GeocodeCache.direccion == direccion_vieja_norm
        ).delete()
    if nueva_norm and nueva_norm != direccion_vieja_norm:
        borradas += db.query(models.GeocodeCache).filter(
            models.GeocodeCache.direccion == nueva_norm
        ).delete()

    db.commit()
    return JSONResponse({
        "ok": True,
        "direccion": nueva,
        "direccion_norm": nueva_norm,
        "cache_borradas": borradas,
    })


@router.get("/mapa-no-ubicadas")
async def mapa_no_ubicadas(request: Request, db: Session = Depends(get_db)):
    """Devuelve la lista de socios cuya dirección no se pudo geocodificar
    (GeocodeCache.lat IS NULL). Incluye datos para mostrarlos en el panel
    "No ubicadas" del mapa con link a editar el socio.
    """
    user = await auth_module.require_user(request, db)
    if not auth_module.has_permission(user, 'compradores', 'ver'):
        raise HTTPException(403, 'No tenés permiso para ver esta sección')

    # Direcciones marcadas como no ubicables en el cache
    fallidas = db.query(models.GeocodeCache.direccion).filter(
        models.GeocodeCache.lat.is_(None)
    ).all()
    fallidas_set = {d for (d,) in fallidas}

    if not fallidas_set:
        return JSONResponse({"ok": True, "items": []})

    rows = (
        db.query(
            models.Boleta, models.Comprador, models.Talonera, models.Vendedor,
        )
        .join(models.Comprador, models.Boleta.comprador_id == models.Comprador.id)
        .join(models.Talonera, models.Boleta.talonera_id == models.Talonera.id)
        .outerjoin(models.Vendedor, models.Boleta.vendedor_id == models.Vendedor.id)
        .filter(models.Comprador.direccion.isnot(None))
        .filter(models.Comprador.direccion != "")
        .filter(models.Talonera.tipo == "COMUN")
        .order_by(models.Comprador.apellido_nombre)
        .all()
    )

    items = []
    for b, c, t, v in rows:
        norm = _norm_direccion(c.direccion or "")
        if norm not in fallidas_set:
            continue
        try:
            num_digitos = t.num_digitos or 4
        except Exception:
            num_digitos = 4
        items.append({
            "comprador_id": c.id,
            "apellido_nombre": c.apellido_nombre,
            "direccion": c.direccion,
            "numero": str(b.numero_principal).zfill(num_digitos),
            "pata": t.nombre,
            "pata_label": (t.nombre or "").replace("PATA ", "X"),
            "vendedor_nombre": v.nombre if v else None,
        })

    return JSONResponse({"ok": True, "items": items, "total": len(items)})
